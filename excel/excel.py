from collections import defaultdict
import openpyxl


class Xl:

    all = defaultdict(list)

    def __init__(self, *args, **kwargs):
        self.object = [*args]
        item = self.search(kwargs.get('filename'), self.object[11])
        if item:
            if self.object[15] > item.object[15]:
                item.object[15] = self.object[15]
            del self
        else:
            self.all[kwargs.get('filename')].append(self)

    def __str__(self):
        return str(self.object[11])

    def __repr__(self):
        return str(self.object[11])

    def search(self, filename, attr):
        it = self.all[filename]
        for item in it:
            if item.object[11] == attr:
                return item
        return None

    @classmethod
    def print_all(cls):
        for key in Xl.all.keys():
            print(key, end='\n')
            for item in Xl.all[key]:
                print('\t', item.object[11], '---', item.object[15])

    @classmethod
    def equal(cls):
        old_list = []
        new_list = []
        first, second = cls.all.keys()
        for item in cls.all[second]:
            old_item = item.search(first, item.object[11])
            if old_item:
                old_item.object[15] = item.object[15]
                old_list.append(old_item)
            else:
                new_list.append(item)
        last_product_id = cls.all[first][-1].object[0]
        for item in new_list:
            last_product_id += 1
            item.object[0] = last_product_id
        return old_list+new_list

    @classmethod
    def load(cls, *args):
        for arg in args:
            file = openpyxl.load_workbook(filename=arg)
            for item in file['Products']:
                if item[0].value and item[0].value != 'product_id':
                    Xl(*(cell.value for cell in item), filename=arg)

    @classmethod
    def save(cls, object_list):
        wb = openpyxl.workbook.Workbook()
        ws = wb.worksheets[0]
        for row in range(1, len(object_list)+1):
            for column in range(1, 42):
                ws.cell(row=row, column=column).value = object_list[row-1].object[column-1]
        print(tuple(ws.rows))
        wb.save('output.xlsx')


Xl.load("products-new.xlsx", "products-new1.xlsx")
edited_list = Xl.equal()
Xl.save(edited_list)